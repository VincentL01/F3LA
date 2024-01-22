import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns

from pathlib import Path

from scipy.stats import pearsonr



class GilbertFileReader():

    def __init__(self, excel_path, given_matcher=None):
        self.wb = openpyxl.load_workbook(excel_path)
        self.sheet_names = self.wb.sheetnames
        self.set_batch_matcher(given_matcher = given_matcher)

    def set_batch_matcher(self, given_matcher=None):
        if given_matcher is not None:
            self.BATCH_MATCHER = given_matcher
        else:
            self.BATCH_MATCHER = {
            "1": ["B5", "P10"],
            "2": ["B12", "P17"],
            "3": ["B19", "P24"],
            "4": ["B26", "P31"],
            "5": ["B33", "P38"],
            "6": ["B40", "P45"],
            }

    def get_column_names(self, sheet_num, columns = ["C", "P"], row = [1, 2]):
        print(f"Getting column names from {sheet_num=} {columns=} {row=}")
        sheet_name = self.sheet_names[sheet_num]
        ws = self.wb[sheet_name]
        # column_names = ['Fish']
        column_names = []
        # iterate from columns[0] to columns[1]
        columns = [chr(ord(columns[0]) + i) for i in range(ord(columns[1]) - ord(columns[0]) + 1)]
        for col in columns:
            _name = ""
            for r in row:
                row_value = ws[col + str(r)].value
                if row_value is not None:
                    _name += row_value + " "
            _name = _name.strip()
            if len(_name) > 0:
                column_names.append(_name)
        
        print("Column names: ", column_names)

        return column_names
    
    def get_data(self, sheet_num, Corner1, Corner2):
        sheet_name = self.sheet_names[sheet_num]
        print(f"Got data from {sheet_name=}, from {Corner1=} to {Corner2=}")
        ws = self.wb[sheet_name]
        df = pd.DataFrame()
        for row in ws[Corner1:Corner2]:
            df[row[0].value] = [cell.value for cell in row[1:]]

        # return transposed dataframe
        transposed_df = df.T
        # remove index
        # transposed_df.reset_index(drop=True, inplace=True)
        # set column names
        new_column_names = self.get_column_names(sheet_num, columns = [Corner1[0], Corner2[0]])

        try:
            transposed_df.columns = new_column_names
        except Exception as e:
            print("df has {} columns, but {} column names were provided.".format(len(transposed_df.columns), len(new_column_names)))
            print(f"Exception: {e}")
            print(f"Transposed DF: {transposed_df}")
            print(new_column_names)

        return transposed_df
    
    def get_info(self, treatment_char, batch_num):
        # change A to 0
        sheet_num = ord(treatment_char) - ord("A")

        Corner1 = self.BATCH_MATCHER[str(batch_num)][0]
        Corner2 = self.BATCH_MATCHER[str(batch_num)][1]

        return self.get_data(sheet_num, Corner1, Corner2)