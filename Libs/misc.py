from pathlib import Path
import math
import json
import pandas as pd
import re
import os
import shutil
from scipy.spatial import ConvexHull
import numpy as np
import openpyxl
import subprocess
import cv2 

import logging

logger = logging.getLogger(__name__)

from . import HISTORY_PATH, EPSILON

def num_to_ord(input_number):
    suf = lambda n: "%d%s"%(n,{1:"st",2:"nd",3:"rd"}.get(n%100 if (n%100)<20 else n%10,"th"))
    if isinstance(input_number, list):
        return [suf(i) for i in input_number]
    elif isinstance(input_number, int):
        return suf(input_number)
    else:
        raise TypeError("Input must be an integer or a list of integers")
    
def ord_to_num(input_ord):
    if isinstance(input_ord, list):
        return [int(re.findall(r'\d+', i)[0]) for i in input_ord]
    elif isinstance(input_ord, str):
        return int(re.findall(r'\d+', input_ord)[0])
    else:
        raise TypeError("Input must be a string or a list of strings")
    
def index_to_char(input_number): # turn 0 to A
    if isinstance(input_number, list):
        return [chr(i+65) for i in input_number]
    elif isinstance(input_number, int):
        return chr(input_number+65)
    else:
        raise TypeError("Input must be an integer or a list of integers")
    
def char_to_index(input_char): # turn A to 0
    if isinstance(input_char, list):
        return [ord(i)-65 for i in input_char]
    elif isinstance(input_char, str):
        return ord(input_char)-65

def hyploader(hyp_path):

    with open(hyp_path, 'r') as file:
        data = json.load(file)
        
    # convert values to int or float
    for key, value in data.items():
        if key == "CONVERSION RATE":
            data[key] = float(value)
        elif key in ["FRAME RATE", "DURATION", "SEGMENT DURATION", "ZONE WIDTH"]:
            data[key] = int(float(value))
        else:
            for fish_num, fish_data in value.items():
                if isinstance(fish_data, list):
                    for i, item in enumerate(fish_data):
                        fish_data[i] = int(float(item))
                else:
                    data[key][fish_num] = int(float(fish_data))

    def zone_calculator(target_name, target_data, reverse=False):
        zone_name = target_name + " ZONE"
        target_data[zone_name] = {}
        n = 1 if reverse == False else -1
        for fish_num, fish_data in target_data[target_name].items():
            m = -1 if fish_data[1] == 0 else 1
            target_data[zone_name][fish_num] = [fish_data[0] + m * n * target_data["ZONE WIDTH"] *  target_data["CONVERSION RATE"], fish_data[1]]

        return target_data

    if "MIRROR" in data.keys():
        data = zone_calculator("MIRROR", data, reverse=True)
    elif "SEPARATOR" in data.keys():
        data = zone_calculator("SEPARATOR", data)

    return data

def calculate_distance(start_point, end_point):
    return math.sqrt((start_point[0]-end_point[0])**2 + (start_point[1]-end_point[1])**2)


def pearson_corr(list1, list2):
    arr1 = np.array(list1)
    arr2 = np.array(list2)
    if len(arr1) == len(arr2):
        return np.corrcoef(arr1, arr2)[0, 1]
    else:
        return "The lists have different lengths!"

def calculate_turning_angle(x1, y1, x2, y2, x3, y3):
    """
    Compute the turning angle (in degrees) at point B=(x2, y2) for a path defined by points A=(x1, y1), B=(x2, y2), and C=(x3, y3).
    The turning angle is the angle between vectors AB and BC.
    Positive values represent right turns, negative values represent left turns.
    """

    # Calculate direction vectors
    D1 = (x2 - x1, y2 - y1)  # BA'
    D2 = (x3 - x2, y3 - y2)  # BC

    # Calculate dot product
    dot_product = D1[0]*D2[0] + D1[1]*D2[1]

    # Calculate magnitudes
    mag_D1 = math.sqrt(D1[0]**2 + D1[1]**2)
    mag_D2 = math.sqrt(D2[0]**2 + D2[1]**2)

    # Calculate cosine of the angle
    try:
        cos_theta = dot_product / (mag_D1 * mag_D2)
    except ZeroDivisionError:
        cos_theta = 0
    cos_theta = max(-1, min(1, cos_theta))
    
    # Calculate the angle in radians
    theta_rad = math.acos(cos_theta)

    # Convert the angle to degrees
    theta_deg = math.degrees(theta_rad)

    # Calculate cross product
    cross_product = D1[0]*D2[1] - D1[1]*D2[0]
    
    # If cross product is positive, make the angle negative
    if cross_product > 0:
        theta_deg = -theta_deg

    return theta_deg


def compute_turning_angle(x1, y1, x2, y2, x3, y3):
    """
    Compute the turning angle (in degrees) at point B=(x2, y2) for a path defined by points A=(x1, y1), B=(x2, y2), and C=(x3, y3).
    The turning angle is the angle between vectors AB and BC.
    Positive values represent right turns, negative values represent left turns.
    """
    dx1 = x2 - x1
    dy1 = y2 - y1
    dx2 = x3 - x2
    dy2 = y3 - y2

    angle1 = np.arctan2(dy1, dx1)
    angle2 = np.arctan2(dy2, dx2)

    return np.degrees((angle1 - angle2 + np.pi) % (2*np.pi) - np.pi)

# def index_filler(index_number, digits=2):
#     index_number = str(index_number)
#     if len(index_number) < digits:
#         return "0"*(digits-len(index_number)) + index_number
#     else:
#         return index_number
    
def event_extractor(binary_list, positive_token = None):

    # find unique values in the binary list
    # binary_list is a list
    unique_values = list(set(binary_list))

    if len(unique_values) == 2:
        if positive_token is None:
            # positive_token = non zero value in unique_values
            positive_token = [i for i in unique_values if i != 0][0]
        else:
            if positive_token not in unique_values:
                raise ValueError("The specified positive token is not in the binary list.")
    elif len(unique_values) > 2:
        if positive_token is None:
            raise ValueError("The binary list has more than two unique values. Please specify the positive token.")
        else:
            if positive_token not in unique_values:
                raise ValueError("The specified positive token is not in the binary list.")
    
    binary_list = [1 if i == positive_token else 0 for i in binary_list]

    result = {}
    start, end = None, None

    for i in range(len(binary_list)):
        if binary_list[i] == 1:
            if start is None:
                start = i
            end = i
        elif start is not None:
            result[(start, end)] = end - start + 1
            start, end = None, None

    if start is not None:
        result[(start, end)] = end - start + 1

    return result

def has_csv_file(directory_path):
    directory = Path(directory_path)
    csv_files = directory.glob('*.csv')

    if len(list(csv_files)) > 0:
        return True
    else:
        return False
    
    
def check_sheet_existence(file_path, sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if sheet_name in workbook.sheetnames:
        return True
    else:
        return False
    
def remove_sheet_by_name(file_path, sheet_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
        workbook.save(filename=file_path)
        return True
    else:
        return False

############################################# FD and Entropy Calculator #############################################

def FD_Entropy_Calculator(input_df):

    def countif(input_list, threshold, less=True):
        count = 0
        for value in input_list:
            if less and value < threshold:
                count+=1
            elif not less and value > threshold:
                count+=1
            
        return count

    x_list = input_df['X']
    y_list = input_df['Y']
    z_list = input_df['Z']

    FRAMES = len(x_list)
    
    times = [i/50 for i in range(FRAMES)] # D

    delta_x = {} #E
    delta_y = {} #F
    delta_z = {} #G
    delta_r = {} #H
    thetas = {}  #I  

    for i in range(FRAMES):
        if i==0:
            continue
        temp_x = x_list[i] - x_list[i-1]
        temp_y = y_list[i] - y_list[i-1]
        temp_z = z_list[i] - z_list[i-1]
        temp_r = math.sqrt(temp_x**2+temp_y**2+temp_z**2)

        delta_x[i] = temp_x
        delta_y[i] = temp_y
        delta_z[i] = temp_z
        delta_r[i] = temp_r

        if i>1:
            dot_product = (delta_x[i]*delta_x[i-1] + delta_y[i]*delta_y[i-1] + delta_z[i]*delta_z[i-1])
            product_of_magnitudes = (delta_r[i]*delta_r[i-1])
            value = dot_product / (product_of_magnitudes + EPSILON)
            value = max(-1, min(1, value))
            temp_theta = math.acos(value)*180/math.pi
            thetas[i] = temp_theta

    # r          #J
    N1 = {}      #K
    # N2 = {       #L
    Cr = {}      #M
    sigma = {}   #N
    logr = {}    #O
    logCr = {}   #P
    
    thresholds = [(i+1)/10 for i in range(FRAMES)]
    index_1 = thresholds.index(1)
    thresholds = thresholds[:index_1] + [1.01] + thresholds[(index_1+1):]
    thresholds[:20]

    for i in range(FRAMES):
        N1[i] = countif(list(delta_r.values()), thresholds[i])
        Cr[i] = N1[i] / (FRAMES - 1)

        try:
            logCr[i] = math.log10(Cr[i])
        except Exception as e:
            logger.error(f"At i = {i}, Cr[i] = {Cr[i]}")
            logger.error(f"Set Cr[i] to 10e-10")
            Cr[i] = 10e-10
            logCr[i] = math.log10(Cr[i])

        sigma[i] = math.log10(Cr[i])/math.log10(thresholds[i])
        logr[i] = math.log10(thresholds[i])

        

    APPROCH = 0.000
    neg_close = math.inf*(-1)
    neg_close_pos = -1
    for i, value in enumerate(list(logr.values())):
        if value < APPROCH and value > neg_close:
            neg_close = value
            neg_close_pos = i

    table_index = list(range(-5, 6))
    table_index.reverse()
    table_index

    FD_df = pd.DataFrame(columns = ['number', 'logari', 'logariC', 'x-xbar', 'y-ybar', 
                            '(x-xbar)(y-ybar)', '(x-xbar)2', '(y-ybar)2', '[yi-(a+bxi)]2'])

    FD_df['number'] = table_index
    LEN = len(table_index)

    # row, col
    for row in range(LEN):
        num = table_index[row]
        num = num + neg_close_pos
        FD_df.iloc[row,1] = logr[num]                             # R
        FD_df.iloc[row,2] = logCr[num]                            # S

    col1 = np.array(list(FD_df['logari']))
    col2 = np.array(list(FD_df['logariC']))
    col1_avg = np.average(col1)                                     # average of R
    col2_avg = np.average(col2)                                     # average of S

    for row in range(LEN):
        FD_df.iloc[row,3] = FD_df.iloc[row,1] - col1_avg             # T
        FD_df.iloc[row,4] = FD_df.iloc[row,2] - col2_avg             # U
        FD_df.iloc[row,5] = FD_df.iloc[row,3] * FD_df.iloc[row,4]    # V
        FD_df.iloc[row,6] = FD_df.iloc[row,3]**2                     # W
        FD_df.iloc[row,7] = FD_df.iloc[row,4]**2                     # X

    #P4 =SUM(U16:U26)/SUM(V16:V26)
    #P5 =AVERAGE(Q16:Q26)-P4*AVERAGE(P16:P26)

    variable_b = np.sum(np.array(list(FD_df['(x-xbar)(y-ybar)']))) / np.sum(np.array(list(FD_df['(x-xbar)2'])))     # P4
    variable_a = col2_avg - col1_avg*variable_b                                                                     # P5

    # '[yi-(a+bxi)]2' =(Q16-($P$5+$P$4*P16))^2
    for row in range(LEN):
            FD_df.iloc[row,8] = (FD_df.iloc[row,2]-(variable_a+variable_b*FD_df.iloc[row,1]))**2     # Y
        
    #P6 =SQRT(SUM(X16:X26)/(COUNT(X16:X26)-2))
    variable_s = np.sqrt( np.sum(np.array(list(FD_df['[yi-(a+bxi)]2']))) / (LEN-2) )

    #R4 =P6/SQRT(SUM(V16:V26))
    variable_bErr = variable_s / np.sqrt(np.sum(np.array(list(FD_df['(x-xbar)2'])))) 

    #R5 =P6*SQRT((1/COUNT(X16:X26))+(AVERAGE(P16:P26)^2)/SUM(V16:V26))
    variable_aErr = variable_s*np.sqrt((1/LEN)+col1_avg**2/np.sum(np.array(list(FD_df['(x-xbar)2']))))

    #RR = =(SUM(U16:U26)^2)/(SUM(V16:V26)*SUM(W16:W26))
    variable_RR = np.sum(np.array(list(FD_df['(x-xbar)(y-ybar)'])))**2 / (np.sum(np.array(list(FD_df['(x-xbar)2']))) * np.sum(np.array(list(FD_df['(y-ybar)2']))))

    def get_entropy():
        G_array = np.array(list(thetas.values()))
        G_count = (G_array >= 90).sum()
        G_count2 = (G_array < 90).sum()
        G_len = G_array.size

        result = (-1) * G_count/G_len * np.log2(G_count/G_len) - G_count2/G_len * np.log2(G_count2/G_len)

        return result

    #H (Entropy)
    variable_Entropy = get_entropy()

    FractalDimension = variable_b
    Entropy = variable_Entropy

    return FractalDimension, Entropy


############################################## SHOALING AREA / VOLUME ##############################################


def HullVolumeCalculator(fishes_coords, surface = ['X', 'Y', 'Z'], save_dir = None):
    volumes = []

    # Assuming that each fish dataframe has the same number of frames
    num_frames = list(fishes_coords.values())[0].shape[0]

    for frame in range(num_frames):
        frame_coords = pd.concat([df.iloc[[frame]] for df in fishes_coords.values()])
        coords_array = frame_coords[surface].to_numpy()
        
        # Calculate the convex hull
        try:
            hull = ConvexHull(coords_array)
            hull_volume = hull.volume
        except: # In case the convex hull cannot be calculated
            hull_volume = 0
        volumes.append(hull_volume)

    # Create a dataframe with the calculated volumes
    df_volumes = pd.DataFrame(volumes, columns=['ConvexHullVolume'])
    df_volumes.index.name = 'Frame'

    if save_dir:               
        save_path = os.path.join(save_dir, f"HullVolume_{len(surface)}D.csv")
        df_volumes.to_csv(save_path, index=True)
        print(f"Saved to {save_path}")


    return df_volumes


############################################## INHERITED FROM OLD CODE ##############################################

def load_raw_df(txt_path, sep = "\t"):
    # Read the .txt file into a dataframe
    raw_df = pd.read_csv(txt_path, sep = sep)
    
    # partial_path = Path(*Path(txt_path).parts[-3:]) if len(Path(txt_path).parts) > 3 else Path(txt_path) 
    
    # if len(raw_df) > 0:
    #     print(f'Loaded raw data from ".\{partial_path}"')
    # else:
    #     print(f'No data loaded from ".\{partial_path}"')

    # Remove unnecessary columns
    tanks_list = []
    for col in raw_df.columns:
        if "unnamed" in col.lower() or "prob" in col.lower():
            raw_df.drop(col, axis=1, inplace=True)
        if "x" in col.lower():
            # find the number in the column name
            tank_num = re.findall(r'\d+', col)
            tank_num = int(tank_num[0])
            tanks_list.append(tank_num)
    # print('Tanks found: ', tanks_list)

    return raw_df, tanks_list

def remove_first_row_if_nan(input_df, limitation):
    # if the first row of the dataframe has nan values, remove the entire first row
    removed_rows = 0
    while len(input_df) > limitation:
        row_0 = input_df.iloc[0, :]
        if row_0.isnull().values.any():
            input_df = input_df.iloc[1:, :]
            removed_rows += 1
            # reset index
            input_df = input_df.reset_index(drop=True)
        else:
            break
    return input_df, removed_rows

def clean_df(input_df, fill = False, frames = 0, remove_nan = True, limitation = 15000): 

    # Remove the initial rows with nan values
    if remove_nan:
        input_df, removed_rows = remove_first_row_if_nan(input_df, limitation)

    # Only take the first frames rows
    if frames == 0:
        pass
    else:
        input_df = input_df.iloc[:frames, :]

    if fill == False:
        return input_df
    
    # Fill the nan values using forward fill and backward fill
    output_df = input_df.fillna(method='ffill')
    output_df = output_df.fillna(method='bfill')

    return output_df, removed_rows

def couple_nan_remover(input_df1, input_df2, limitation = 15000):
    # Balance 2 dataframes
    
    length_diff = len(input_df1) - len(input_df2)
    # Remove the last length_diff rows of the longer dataframe
    if length_diff > 0:
        input_df1 = input_df1.iloc[:-length_diff, :]
        logger.debug(f"Removed {length_diff} rows from the end of input_df1")
    elif length_diff < 0:
        input_df2 = input_df2.iloc[:length_diff, :]
        logger.debug(f"Removed {length_diff} rows from the end of input_df2")

    remove_window = len(input_df1) - limitation
    origin_length = len(input_df1)
    drop_rows = []

    for row_num in range(origin_length):
        try:
            df1_row = input_df1.iloc[row_num, :]
            df2_row = input_df2.iloc[row_num, :]
        except:
            break

        if df1_row.isnull().values.any() or df2_row.isnull().values.any():
            drop_rows.append(row_num)
            remove_window -= 1

            # reset index
            input_df1 = input_df1.reset_index(drop=True)
            input_df2 = input_df2.reset_index(drop=True)

        if remove_window <= 0:
            break

    # Remove the row in drop_rows from the dataframe:
    input_df1.drop(drop_rows)
    logger.debug(f"Removed {len(drop_rows)} nan-containing rows from first dataframe")
    input_df2.drop(drop_rows)
    logger.debug(f"Removed {len(drop_rows)} nan-containing rows from second dataframe")

    if len(input_df1) > limitation:
        logger.debug(f"Dataframe have {len(input_df1)=} rows, only take the first {limitation} rows")
        input_df1 = input_df1.iloc[:limitation, :]
        input_df2 = input_df2.iloc[:limitation, :]

    return input_df1, input_df2


def couple_df_cleaner(input_df1, input_df2, fill = True, remove_nan = True, limitation = 15000):
    # Remove the initial rows with nan values
    if remove_nan:
        logger.info("Removing nans from the beginning of two given trajectories dataframes")
        input_df1, input_df2 = couple_nan_remover(input_df1, input_df2, limitation)

    if fill == False:
        return input_df1, input_df2
    
    # Fill the nan values using forward fill and backward fill
    output_df1 = input_df1.fillna(method='ffill')
    output_df1 = output_df1.fillna(method='bfill')

    output_df2 = input_df2.fillna(method='ffill')
    output_df2 = output_df2.fillna(method='bfill')

    # Double check if output_df1 has nan values
    if output_df1.isnull().values.any():
        logger.warning("First dataframe still has nan values")
        raise ValueError("First dataframe still has nan values")
    else:
        logger.info("First dataframe is clean of nan values")

    # Double check if output_df2 has nan values
    if output_df2.isnull().values.any():
        logger.warning("Second dataframe still has nan values")
        raise ValueError("Second dataframe still has nan values")
    else:
        logger.info("First dataframe is clean of nan values")

    return output_df1, output_df2

def append_df_to_excel(filename, df, sheet_name='Sheet1', startcol=None, startrow=None, col_sep = 0, row_sep = 0,
                       truncate_sheet=False, DISPLAY = False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        try:
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startcol=startcol if startcol is not None else 0, 
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            logger.info(f"Successful write to {filename}, sheet={sheet_name} at column = {startcol}, row = {startrow}")
        except Exception as e:
            logger.warning(f"EXCEPTION: {e}")
            logger.warning(f"UNSUCCESS write to {filename}, sheet={sheet_name} at column = {startcol}, row = {startrow}")
            logger.debug(f"df: {df}")

        # wb = openpyxl.load_workbook(filename)
        # ws = wb[sheet_name]
        # row_0 = ws[1]
        return
    

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.workbook = openpyxl.load_workbook(filename)

    # get the last col in the existing Excel sheet
    # if it was not specified explicitly
    if startcol is None and sheet_name in writer.workbook.sheetnames:
        startcol = writer.workbook[sheet_name].max_column + col_sep

    if startrow is None and sheet_name in writer.workbook.sheetnames:
        startrow = writer.workbook[sheet_name].max_row + row_sep
    
    if startcol is None:
        startcol = 0

    if startrow is None:
        startrow = 0
    
    # row_0 = writer.workbook[sheet_name][1]
    # logger.debug(f"Header: {row_0}")
    
    # remove df headers if they exist
    if startrow != 0:
        # take the first row
        first_row = df.iloc[0].astype(str)
        # check if any cell in the first row contains a letter
        has_letter = first_row.str.contains('[a-zA-Z]').any()
        if has_letter:
            df = df.iloc[1:, :]

    # write the dataframe to the existing sheet
    try:
        df.to_excel(writer, sheet_name, startcol=startcol, startrow=startrow, **to_excel_kwargs)
        logger.info(f"Successful write to {filename}/{sheet_name} at column = {startcol}, row = {startrow}")
    except:
        logger.warning(f"UNSUCCESS write to {filename}/{sheet_name} at column = {startcol}, row = {startrow}")

    # close workbook
    writer.close()


def merge_cells(file_path, input_sheet_name = None, input_column_name = 'Shoaling Area', cell_step=3, inplace = True):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)

    if input_sheet_name == None:
        sheet_names = workbook.worksheets
    elif isinstance(input_sheet_name, list):
        sheet_names = input_sheet_name
    elif isinstance(input_sheet_name, str):
        sheet_names = [input_sheet_name]
    elif isinstance(input_sheet_name, int):
        sheet_names = [workbook.worksheets[input_sheet_name]]

    if isinstance(input_column_name, str):
        column_names = [input_column_name]
    elif isinstance(input_column_name, list):
        column_names = input_column_name

    logger.debug(f"Merging {cell_step} rows of {column_names} in {sheet_names}")

    for worksheet in workbook.worksheets:
        if worksheet not in sheet_names:
            continue
        # Find the column index for the "Shoaling Area" header
        for colume_name in column_names:
            shoaling_area_col = None
            for col_idx in range(1, worksheet.max_column+1):
                header = worksheet.cell(row=1, column=col_idx).value
                if header and colume_name in header:
                    shoaling_area_col = col_idx
                    break

            if shoaling_area_col is None:
                print("Column not found.")
            else:
                # Merge every next 3 rows of the Shoaling Area column
                for row_idx in range(2, worksheet.max_row+1, cell_step):
                    value = worksheet.cell(row=row_idx, column=shoaling_area_col).value
                    # print(value)
                    if value is not None:
                        # Merge the current row with the next 2 rows
                        worksheet.merge_cells(start_row=row_idx, start_column=shoaling_area_col, end_row=row_idx+2, end_column=shoaling_area_col)
                    
                    # align the merged cell, horizontal and vertical center
                    worksheet.cell(row=row_idx, column=shoaling_area_col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
    # define output_path
    if inplace == False:
        output_path = file_path[:-5] + '_merged.xlsx'        
    else:
        output_path = file_path    

    # Save the modified workbook
    try:
        workbook.save(filename=output_path)
        logger.debug(f"Merged {cell_step} rows of {column_names} in {sheet_names}")
    except:
        logger.warning(f"UNSUCCESS merge for {file_path}")


def excel_polish(file_path, batch_num=1, inplace=True):

    logger.debug("Polishing excel file...")

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename=file_path)


    # Adjust the column widths
    # Loop through each sheet in the workbook
    for sheet_name in workbook.sheetnames:

        logger.debug(f"In sheet name: {sheet_name}")

        if "analysis" in sheet_name.lower():
            continue
        # Select the sheet
        sheet = workbook[sheet_name]
        
        # Loop through each column in the sheet
        for col in sheet.columns:
            # Set the width of the column to 17.00 (160 pixels)
            sheet.column_dimensions[col[0].column_letter].width = 17.00

            logger.debug(f"Set column {col[0].column_letter} width to 17.00")
        
        # Enable text wrapping for the header row
        for cell in sheet[1]:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')


    # Save the modified workbook
    try:
        workbook.save(filename=file_path)
        logger.info(f"Polished completely for {file_path}")
    except:
        logger.info(f"UNSUCCESS polish for {file_path}")


def open_explorer(path):
    # Check if the given path exists
    if os.path.exists(path):
        # Open the file explorer
        subprocess.run(['explorer', os.path.realpath(path)])
    else:
        print("The provided path does not exist.")



##################################### CONSTANT GENERATOR #####################################

def get_working_dir(project_dir, batch_num):
    return Path(project_dir) / f"Batch {batch_num}"

def get_treatment_dir(project_dir, batch_num, treatment_char):
    working_dir = get_working_dir(project_dir, batch_num)

    treatment_pattern = f"{treatment_char} - "

    FOUND = False

    for child_dir in os.listdir(working_dir):
        if treatment_pattern in child_dir:
            FOUND = True
            return working_dir / child_dir

    if FOUND == False:
        raise FileNotFoundError(f"Couldn't find treatment directory with pattern [{treatment_char} -]")
    

def get_static_dir(project_dir, batch_num, treatment_char):
    working_dir = get_working_dir(project_dir, batch_num)
    return working_dir / "static" / treatment_char

def get_trajectories_dir(project_dir, batch_num, treatment_char):
    static_dir = get_static_dir(project_dir, batch_num, treatment_char)
    return static_dir / "trajectories"

def get_normalized_trajectories_dir(project_dir, batch_num, treatment_char, unit):
    static_dir = get_static_dir(project_dir, batch_num, treatment_char)
    return static_dir / f"trajectories_normalized_{unit}"

def get_sideview_trajectory_path(project_dir, batch_num, treatment_char):
    treatment_dir = get_treatment_dir(project_dir, batch_num, treatment_char)
    return treatment_dir / "Side View" / "trajectories_nogaps.txt"

def get_topview_trajectory_path(project_dir, batch_num, treatment_char):
    treatment_dir = get_treatment_dir(project_dir, batch_num, treatment_char)
    return treatment_dir / "Top View" / "trajectories_nogaps.txt"


################################################# SORTER ##########################################################

def find_the_whole_num(given_string, end_index):

    num_in_string = ""

    i = end_index

    while True:
        num_in_string = given_string[i] + num_in_string
        try:
            previous_char = given_string[i - 1]
        except IndexError:
            break

        if previous_char.isdigit():
            i -= 1
        else:
            break
            
    return int(num_in_string)


def find_batch_num(given_string):
    # find "st", "nd", "rd", "th" in a string
    # if found, check if the previous character is a number
    # if yes, return the number
    # if no, return None

    # find all occurrences of "st", "nd", "rd", "th"
    occurrences = [m.start() for m in re.finditer('st|nd|rd|th', given_string)]

    num_occurrence = [occ - 1 for occ in occurrences if given_string[occ - 1].isdigit()]

    if len(num_occurrence) != 1:
        return None
    
    num_occurrence = num_occurrence[0]

    return find_the_whole_num(given_string, num_occurrence)


def find_treatment_num(given_string):

    indicator = given_string.split("-")[0].strip()

    # Check if indicator is number of char
    try:
        indicator = int(indicator)
    except ValueError:
        # change A -> 1, B -> 2, C -> 3, etc.
        try:
            indicator = char_to_index(indicator)
        except Exception as e:
            message = f"[STRUCTURE ERROR] The indicator before the dash char ( - ) in Treatment folder name is unusual!\n{e}"
            logger.error(f"{indicator=}")
            raise Exception(message)

    return indicator


def substance_dose_unit_finder(given_string):

    parts = given_string.split()

    if len(parts) == 1:
        return given_string, "", ""
    elif len(parts) >= 2:
        substance = ""
        for part in parts:
            if re.search('[0-9]', part) and re.search('[a-zA-Z]', part):
                unit = re.findall("[a-zA-Z]+", part)[0]
                dose = part.replace(unit, "")
            else:
                substance += part + " "
        return substance.strip(), dose, unit


class Importer():

    def __init__(self, import_project_dir, target_project_dir, trajectories_format="trajectories_nogaps.txt"):
                 
        self.import_project_dir = Path(import_project_dir)
        self.target_project_dir = Path(target_project_dir)
        self.trajectories_format = trajectories_format

        self.new_treatments = []


    def import_trajectories(self):
        import_data = self.data_sorter()
        self.data_distributor(import_data)


    def data_sorter(self):
        # find all directories inside
        treatment_dirs = [x for x in self.import_project_dir.iterdir() if x.is_dir() and "-" in x.name]

        import_data = {}

        self.import_treatment_names = {}

        for treatment_dir in treatment_dirs:
            treatment_char = index_to_char(find_treatment_num(treatment_dir.name)-1)
            import_data[treatment_char] = {}
            self.import_treatment_names[treatment_char] = treatment_dir.name.split("-")[1].split("(")[0].strip()

            batch_dirs = [x for x in treatment_dir.iterdir() if x.is_dir()]

            for batch_dir in batch_dirs:
                batch_name = batch_dir.name
                batch_num = find_batch_num(batch_name)
                if batch_num not in import_data[treatment_char]:
                    import_data[treatment_char][batch_num] = {}
                if "side view" in batch_name.lower():
                    import_data[treatment_char][batch_num]["Side View"] = batch_dir / self.trajectories_format
                elif "top view" in batch_name.lower():
                    import_data[treatment_char][batch_num]["Top View"] = batch_dir / self.trajectories_format

        return import_data


    def data_distributor(self, import_data):
        for treatment_char in import_data:
            for batch_num in import_data[treatment_char]:
                for view in import_data[treatment_char][batch_num]:
                    logger.debug(f"Working with {treatment_char} - Batch {batch_num} - {view}")
                    target_path = self.get_project_path(treatment_char, batch_num, view)
                    if target_path.exists():
                        logger.debug(f"[WARNING] {target_path} already exists!")
                    # copy the file from import_data to target_path
                    shutil.copy(import_data[treatment_char][batch_num][view], target_path)
                    logger.debug(f"Copied {import_data[treatment_char][batch_num][view]} to {target_path}")


    def get_project_path(self, treatment_char, batch_num, view):
        batch_dir = self.target_project_dir / f"Batch {batch_num}"
        # find within batch_dir, folder with f"{treatmentchar} -"
        try:
            treatment_dir = [x for x in batch_dir.iterdir() if x.is_dir() and f"{treatment_char} -" in x.name][0]
        except:
            batch_dir.mkdir(exist_ok=True)
            treatment_dir = batch_dir / f"{treatment_char} - {self.import_treatment_names[treatment_char]}"
            treatment_dir.mkdir(exist_ok=True)
            logger.warning("Treatment folder not found! Creating new folder based on import data...")
            new_info = {
                "char": treatment_char,
                "name": self.import_treatment_names[treatment_char],
                "batch_num": batch_num
            }
            self.new_treatments.append(new_info)

        view_dir = treatment_dir / view
        view_dir.mkdir(exist_ok=True)
        view_path = view_dir / self.trajectories_format

        return view_path
    

def initiator():

    # check if HISTORY_PATH Exists
    if not HISTORY_PATH.exists():
        #make an empty .json history file
        with open(HISTORY_PATH, "w") as f:
            json.dump({}, f, indent=4)

        logger.info("No history file found! Creating new history file...")


def get_first_frame(video_path):
    cap = cv2.VideoCapture(video_path)
    cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
    ret, frame = cap.read()
    cap.release()
    return frame

